#!/usr/bin/env python3
"""
Script principal pour générer toutes les galeries eBird v2.0
Bilingue français/anglais avec support taxonomie

Configuration: config.yaml
"""

import sys
import re
import csv
import unicodedata
from pathlib import Path
from datetime import datetime, timedelta

try:
    import yaml
except ImportError:
    print("❌ PyYAML non installé. Exécutez: pip install pyyaml")
    sys.exit(1)

try:
    from jinja2 import Environment, FileSystemLoader
except ImportError:
    print("❌ Jinja2 non installé. Exécutez: pip install jinja2")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    openpyxl = None  # Optional - only for threatened species

from generate_gallery import (
    EBirdGalleryGenerator,
    verifier_tous_les_medias,
    generer_description_groupe_fr,
    mettre_au_pluriel,
    formater_plage_dates,
    formater_date,
    charger_cache,
    ecrire_feed_json,
    normaliser_nom_commun,
    normaliser_nom_scientifique
)


# ============================================================================
# CHARGEMENT CONFIGURATION
# ============================================================================

CONFIG_FILE = "config.yaml"

def charger_config():
    """Charge la configuration depuis le fichier YAML"""
    if not Path(CONFIG_FILE).exists():
        print(f"❌ Fichier de configuration non trouvé: {CONFIG_FILE}")
        sys.exit(1)
    
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    
    return config


# ============================================================================
# CHARGEMENT STATUTS IUCN
# ============================================================================

def charger_statuts_iucn(config_menacees: dict) -> dict:
    """
    Charge les statuts IUCN depuis un fichier Excel
    
    Returns:
        dict: {nom_scientifique: code_statut}
    """
    if not openpyxl:
        print("   ⚠ openpyxl non installé. Exécutez: pip install openpyxl")
        return {}
    
    fichier = config_menacees.get('fichier_statuts')
    if not fichier or not Path(fichier).exists():
        print(f"   ⚠ Fichier statuts IUCN non trouvé: {fichier}")
        return {}
    
    colonne_nom = config_menacees.get('colonne_nom_scientifique', 'Scientific_name')
    colonne_statut = config_menacees.get('colonne_statut', 'IUCN_Red_List_Category')
    colonne_rang = config_menacees.get('colonne_rang_taxonomique', 'Taxon_rank')
    valeur_espece = config_menacees.get('valeur_rang_espece', 'species')
    
    statuts = {}
    
    try:
        wb = openpyxl.load_workbook(fichier, read_only=True)
        ws = wb.active
        
        # Trouver les indices des colonnes
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        try:
            idx_nom = headers.index(colonne_nom)
            idx_statut = headers.index(colonne_statut)
            idx_rang = headers.index(colonne_rang) if colonne_rang in headers else None
        except ValueError as e:
            print(f"   ⚠ Colonne non trouvée dans {fichier}: {e}")
            return {}
        
        # Lire les données
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Filtrer sur le rang taxonomique si spécifié
            if idx_rang is not None and row[idx_rang] != valeur_espece:
                continue
            
            nom_sci = row[idx_nom]
            statut = row[idx_statut]
            
            if nom_sci and statut:
                statuts[nom_sci] = statut
        
        wb.close()
        print(f"   ✓ Statuts IUCN chargés: {len(statuts)} espèces")
        
    except Exception as e:
        print(f"   ⚠ Erreur lecture {fichier}: {e}")
    
    return statuts


# ============================================================================
# MÉTADONNÉES DES PAYS (drapeaux, coordonnées, continents)
# ============================================================================

COUNTRY_METADATA = {
    # Amérique du Nord
    'CA': {'flag': '🇨🇦', 'lat': 56.13, 'lng': -106.35, 'continent': 'north_america', 'continent_fr': 'Amérique du Nord', 'continent_en': 'North America'},
    'US': {'flag': '🇺🇸', 'lat': 37.09, 'lng': -95.71, 'continent': 'north_america', 'continent_fr': 'Amérique du Nord', 'continent_en': 'North America'},
    'MX': {'flag': '🇲🇽', 'lat': 23.63, 'lng': -102.55, 'continent': 'north_america', 'continent_fr': 'Amérique du Nord', 'continent_en': 'North America'},
    
    # Amérique centrale et Caraïbes
    'CR': {'flag': '🇨🇷', 'lat': 9.75, 'lng': -83.75, 'continent': 'central_america', 'continent_fr': 'Amérique centrale', 'continent_en': 'Central America'},
    'PA': {'flag': '🇵🇦', 'lat': 8.54, 'lng': -80.78, 'continent': 'central_america', 'continent_fr': 'Amérique centrale', 'continent_en': 'Central America'},
    'CU': {'flag': '🇨🇺', 'lat': 21.52, 'lng': -77.78, 'continent': 'central_america', 'continent_fr': 'Caraïbes', 'continent_en': 'Caribbean'},
    'DO': {'flag': '🇩🇴', 'lat': 18.74, 'lng': -70.16, 'continent': 'central_america', 'continent_fr': 'Caraïbes', 'continent_en': 'Caribbean'},
    'JM': {'flag': '🇯🇲', 'lat': 18.11, 'lng': -77.30, 'continent': 'central_america', 'continent_fr': 'Caraïbes', 'continent_en': 'Caribbean'},
    'PR': {'flag': '🇵🇷', 'lat': 18.22, 'lng': -66.59, 'continent': 'central_america', 'continent_fr': 'Caraïbes', 'continent_en': 'Caribbean'},
    'GT': {'flag': '🇬🇹', 'lat': 15.78, 'lng': -90.23, 'continent': 'central_america', 'continent_fr': 'Amérique centrale', 'continent_en': 'Central America'},
    'BZ': {'flag': '🇧🇿', 'lat': 17.19, 'lng': -88.50, 'continent': 'central_america', 'continent_fr': 'Amérique centrale', 'continent_en': 'Central America'},
    'HN': {'flag': '🇭🇳', 'lat': 15.20, 'lng': -86.24, 'continent': 'central_america', 'continent_fr': 'Amérique centrale', 'continent_en': 'Central America'},
    'NI': {'flag': '🇳🇮', 'lat': 12.87, 'lng': -85.21, 'continent': 'central_america', 'continent_fr': 'Amérique centrale', 'continent_en': 'Central America'},
    'SV': {'flag': '🇸🇻', 'lat': 13.79, 'lng': -88.90, 'continent': 'central_america', 'continent_fr': 'Amérique centrale', 'continent_en': 'Central America'},
    
    # Amérique du Sud
    'CL': {'flag': '🇨🇱', 'lat': -35.68, 'lng': -71.54, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    'AR': {'flag': '🇦🇷', 'lat': -38.42, 'lng': -63.62, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    'BR': {'flag': '🇧🇷', 'lat': -14.24, 'lng': -51.93, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    'PE': {'flag': '🇵🇪', 'lat': -9.19, 'lng': -75.02, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    'CO': {'flag': '🇨🇴', 'lat': 4.57, 'lng': -74.30, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    'EC': {'flag': '🇪🇨', 'lat': -1.83, 'lng': -78.18, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    'VE': {'flag': '🇻🇪', 'lat': 6.42, 'lng': -66.59, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    'BO': {'flag': '🇧🇴', 'lat': -16.29, 'lng': -63.59, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    'PY': {'flag': '🇵🇾', 'lat': -23.44, 'lng': -58.44, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    'UY': {'flag': '🇺🇾', 'lat': -32.52, 'lng': -55.77, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    'GY': {'flag': '🇬🇾', 'lat': 4.86, 'lng': -58.93, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    'SR': {'flag': '🇸🇷', 'lat': 3.92, 'lng': -56.03, 'continent': 'south_america', 'continent_fr': 'Amérique du Sud', 'continent_en': 'South America'},
    
    # Europe
    'FR': {'flag': '🇫🇷', 'lat': 46.23, 'lng': 2.21, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'CH': {'flag': '🇨🇭', 'lat': 46.82, 'lng': 8.23, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'IT': {'flag': '🇮🇹', 'lat': 41.87, 'lng': 12.57, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'ES': {'flag': '🇪🇸', 'lat': 40.46, 'lng': -3.75, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'PT': {'flag': '🇵🇹', 'lat': 39.40, 'lng': -8.22, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'GB': {'flag': '🇬🇧', 'lat': 55.38, 'lng': -3.44, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'DE': {'flag': '🇩🇪', 'lat': 51.17, 'lng': 10.45, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'AT': {'flag': '🇦🇹', 'lat': 47.52, 'lng': 14.55, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'NL': {'flag': '🇳🇱', 'lat': 52.13, 'lng': 5.29, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'BE': {'flag': '🇧🇪', 'lat': 50.50, 'lng': 4.47, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'IE': {'flag': '🇮🇪', 'lat': 53.14, 'lng': -7.69, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'GR': {'flag': '🇬🇷', 'lat': 39.07, 'lng': 21.82, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'TR': {'flag': '🇹🇷', 'lat': 38.96, 'lng': 35.24, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'NO': {'flag': '🇳🇴', 'lat': 60.47, 'lng': 8.47, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'SE': {'flag': '🇸🇪', 'lat': 60.13, 'lng': 18.64, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'FI': {'flag': '🇫🇮', 'lat': 61.92, 'lng': 25.75, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'DK': {'flag': '🇩🇰', 'lat': 56.26, 'lng': 9.50, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'IS': {'flag': '🇮🇸', 'lat': 64.96, 'lng': -19.02, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'PL': {'flag': '🇵🇱', 'lat': 51.92, 'lng': 19.15, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'CZ': {'flag': '🇨🇿', 'lat': 49.82, 'lng': 15.47, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'HU': {'flag': '🇭🇺', 'lat': 47.16, 'lng': 19.50, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'RO': {'flag': '🇷🇴', 'lat': 45.94, 'lng': 24.97, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'HR': {'flag': '🇭🇷', 'lat': 45.10, 'lng': 15.20, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    'SI': {'flag': '🇸🇮', 'lat': 46.15, 'lng': 14.99, 'continent': 'europe', 'continent_fr': 'Europe', 'continent_en': 'Europe'},
    
    # Asie
    'JP': {'flag': '🇯🇵', 'lat': 36.20, 'lng': 138.25, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'CN': {'flag': '🇨🇳', 'lat': 35.86, 'lng': 104.20, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'IN': {'flag': '🇮🇳', 'lat': 20.59, 'lng': 78.96, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'TH': {'flag': '🇹🇭', 'lat': 15.87, 'lng': 100.99, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'VN': {'flag': '🇻🇳', 'lat': 14.06, 'lng': 108.28, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'ID': {'flag': '🇮🇩', 'lat': -0.79, 'lng': 113.92, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'MY': {'flag': '🇲🇾', 'lat': 4.21, 'lng': 101.98, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'SG': {'flag': '🇸🇬', 'lat': 1.35, 'lng': 103.82, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'PH': {'flag': '🇵🇭', 'lat': 12.88, 'lng': 121.77, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'KR': {'flag': '🇰🇷', 'lat': 35.91, 'lng': 127.77, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'TW': {'flag': '🇹🇼', 'lat': 23.70, 'lng': 120.96, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'IL': {'flag': '🇮🇱', 'lat': 31.05, 'lng': 34.85, 'continent': 'asia', 'continent_fr': 'Moyen-Orient', 'continent_en': 'Middle East'},
    'AE': {'flag': '🇦🇪', 'lat': 23.42, 'lng': 53.85, 'continent': 'asia', 'continent_fr': 'Moyen-Orient', 'continent_en': 'Middle East'},
    'NP': {'flag': '🇳🇵', 'lat': 28.39, 'lng': 84.12, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    'LK': {'flag': '🇱🇰', 'lat': 7.87, 'lng': 80.77, 'continent': 'asia', 'continent_fr': 'Asie', 'continent_en': 'Asia'},
    
    # Afrique
    'ZA': {'flag': '🇿🇦', 'lat': -30.56, 'lng': 22.94, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'KE': {'flag': '🇰🇪', 'lat': -0.02, 'lng': 37.91, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'TZ': {'flag': '🇹🇿', 'lat': -6.37, 'lng': 34.89, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'MA': {'flag': '🇲🇦', 'lat': 31.79, 'lng': -7.09, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'EG': {'flag': '🇪🇬', 'lat': 26.82, 'lng': 30.80, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'CI': {'flag': '🇨🇮', 'lat': 7.54, 'lng': -5.55, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'SN': {'flag': '🇸🇳', 'lat': 14.50, 'lng': -14.45, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'GH': {'flag': '🇬🇭', 'lat': 7.95, 'lng': -1.02, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'CM': {'flag': '🇨🇲', 'lat': 7.37, 'lng': 12.35, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'NA': {'flag': '🇳🇦', 'lat': -22.96, 'lng': 18.49, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'BW': {'flag': '🇧🇼', 'lat': -22.33, 'lng': 24.68, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'UG': {'flag': '🇺🇬', 'lat': 1.37, 'lng': 32.29, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'RW': {'flag': '🇷🇼', 'lat': -1.94, 'lng': 29.87, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'ET': {'flag': '🇪🇹', 'lat': 9.15, 'lng': 40.49, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    'MG': {'flag': '🇲🇬', 'lat': -18.77, 'lng': 46.87, 'continent': 'africa', 'continent_fr': 'Afrique', 'continent_en': 'Africa'},
    
    # Océanie
    'AU': {'flag': '🇦🇺', 'lat': -25.27, 'lng': 133.78, 'continent': 'oceania', 'continent_fr': 'Océanie', 'continent_en': 'Oceania'},
    'NZ': {'flag': '🇳🇿', 'lat': -40.90, 'lng': 174.89, 'continent': 'oceania', 'continent_fr': 'Océanie', 'continent_en': 'Oceania'},
    'FJ': {'flag': '🇫🇯', 'lat': -17.71, 'lng': 178.07, 'continent': 'oceania', 'continent_fr': 'Océanie', 'continent_en': 'Oceania'},
    'PG': {'flag': '🇵🇬', 'lat': -6.31, 'lng': 143.96, 'continent': 'oceania', 'continent_fr': 'Océanie', 'continent_en': 'Oceania'},
}


# ============================================================================
# CONSTRUCTION DU MENU
# ============================================================================

def construire_menu(config, generator, curation=None):
    """Construit la structure du menu de navigation"""
    
    # Récupérer les familles avec photos
    familles_avec_photos = generator.get_families_with_photos()
    
    # Vérifier s'il y a des photos best
    has_best_photos = curation and any(status == 'best' for status in curation.values())
    
    menu = [
        {
            'name_fr': 'Accueil',
            'name_en': 'Home',
            'file_fr': config['site'].get('index_fr', 'index_fr.html'),
            'file_en': config['site'].get('index_en', 'index_en.html')
        },
        {
            'name_fr': 'Espèces',
            'name_en': 'Species',
            'file_fr': 'species_list_fr.html',
            'file_en': 'species_list_en.html'
        },
        {
            'name_fr': 'Ajouts récents',
            'name_en': 'Recent Updates',
            'file_fr': 'gallery_recent_fr.html',
            'file_en': 'gallery_recent_en.html'
        }
    ]
    
    # Lien Meilleures photos (si des photos best existent)
    if has_best_photos:
        menu.append({
            'name_fr': 'Meilleures photos',
            'name_en': 'Best of',
            'file_fr': 'gallery_best_fr.html',
            'file_en': 'gallery_best_en.html'
        })
    
    # Lien Voyages (si configurés)
    voyages = config.get('voyages', [])
    if voyages:
        menu.append({
            'name_fr': 'Voyages',
            'name_en': 'Trips',
            'file_fr': 'voyages_index_fr.html',
            'file_en': 'voyages_index_en.html'
        })
    
    # Lien Pays (toujours présent si des photos existent)
    menu.append({
        'name_fr': 'Pays',
        'name_en': 'Countries',
        'file_fr': 'pays_index_fr.html',
        'file_en': 'pays_index_en.html'
    })
    
    # Lien Espèces menacées (si configuré)
    config_menacees = config.get('especes_menacees', {})
    if config_menacees.get('activer', False):
        menu.append({
            'name_fr': 'Menacées',
            'name_en': 'Endangered',
            'file_fr': 'menacees_fr.html',
            'file_en': 'menacees_en.html'
        })
    
    # Lien Sons (seulement si le cache médias contient au moins un son)
    has_sounds = any(v.get('status') == 'son' for v in generator.media_cache.values())
    if has_sounds:
        menu.append({
            'name_fr': 'Sons',
            'name_en': 'Sounds',
            'file_fr': 'sounds_gallery_fr.html',
            'file_en': 'sounds_gallery_en.html'
        })
    
    # Construire les données des familles pour la génération
    species_list = generator.get_species_list()
    
    # Grouper les espèces par famille pour les descriptions
    familles_especes = {}
    for sp in species_list:
        fam = sp['family'] or 'Unknown'
        if fam not in familles_especes:
            familles_especes[fam] = []
        familles_especes[fam].append(sp['common_name_fr'])
    
    familles_data = []
    for fam_info in familles_avec_photos:
        fam_code = fam_info['code']
        
        if fam_code == 'Unknown':
            continue
        
        # Générer description française
        noms_fr = familles_especes.get(fam_code, [])
        desc_fr = generer_description_groupe_fr(noms_fr)
        
        # Extraire description anglaise
        name_en_match = re.search(r'\(([^)]+)\)', fam_info['name'])
        desc_en = name_en_match.group(1) if name_en_match else ''
        
        file_id = f"gallery_{fam_code.lower()}"

        familles_data.append({
            'family_code': fam_code,
            'desc_fr': desc_fr,
            'desc_en': desc_en,
            'output_base': file_id,
            'file_fr': f"{file_id}_fr.html",
            'file_en': f"{file_id}_en.html"
        })
    
    return menu, familles_data


# ============================================================================
# HELPERS
# ============================================================================

def obtenir_frontispice(photos: list, curation: dict = None) -> str:
    """
    Obtient le numéro ML du frontispice (image de couverture).
    Priorité: photo 'best' la plus récente > photo la plus récente
    """
    if not photos:
        return None
    
    # Chercher les photos "best"
    if curation:
        best_photos = [p for p in photos if curation.get(p['ml_catalog_number']) == 'best']
        if best_photos:
            # Trier par date décroissante et prendre la plus récente
            best_photos.sort(key=lambda x: x.get('date_raw', ''), reverse=True)
            return best_photos[0]['ml_catalog_number']
    
    # Sinon, prendre la photo la plus récente
    photos_sorted = sorted(photos, key=lambda x: x.get('date_raw', ''), reverse=True)
    return photos_sorted[0]['ml_catalog_number']


# ============================================================================
# GÉNÉRATION
# ============================================================================

def verifier_curation_downloads(curation_file: str):
    """
    Vérifie si le fichier de curation existe dans ~/Downloads et le copie
    dans le répertoire courant si c'est le cas.
    """
    import shutil
    
    # Chemin vers ~/Downloads/photo_curation.csv
    downloads_path = Path.home() / 'Downloads' / Path(curation_file).name
    local_path = Path(curation_file)
    
    if downloads_path.exists():
        # Vérifier si le fichier local existe et comparer les dates
        if local_path.exists():
            downloads_mtime = downloads_path.stat().st_mtime
            local_mtime = local_path.stat().st_mtime
            
            if downloads_mtime > local_mtime:
                # Le fichier dans Downloads est plus récent
                shutil.copy2(downloads_path, local_path)
                print(f"📥 Curation mise à jour depuis ~/Downloads/{downloads_path.name}")
                return True
            else:
                # Le fichier local est déjà à jour
                return False
        else:
            # Pas de fichier local, copier depuis Downloads
            shutil.copy2(downloads_path, local_path)
            print(f"📥 Curation copiée depuis ~/Downloads/{downloads_path.name}")
            return True
    
    return False


def generer_toutes_galeries():
    """Génère toutes les galeries configurées"""
    
    print("=" * 60)
    print("🐦 GÉNÉRATION DES GALERIES eBird v2.0")
    print("=" * 60)
    
    # Charger la configuration
    config = charger_config()
    
    # Extraire les chemins des fichiers
    fichiers = config.get('fichiers', {})
    CSV_FILE = fichiers.get('donnees_ebird', 'MyEBirdData.csv')
    TAXONOMY_FILE = fichiers.get('taxonomie', 'eBird_taxonomy_v2025.csv')
    MEDIA_CACHE = fichiers.get('cache_medias', 'media_cache.csv')
    TRADUCTIONS_FILE = fichiers.get('traductions_lieux', 'traductions_lieux.csv')
    CURATION_FILE = fichiers.get('curation', 'photo_curation.csv')
    
    # Vérifier si un nouveau fichier de curation est dans Downloads
    verifier_curation_downloads(CURATION_FILE)
    
    # Charger la curation
    curation = charger_curation(CURATION_FILE)
    if curation:
        best_count = sum(1 for v in curation.values() if v == 'best')
        reject_count = sum(1 for v in curation.values() if v == 'reject')
        print(f"\n📋 Curation chargée: {len(curation)} photos (⭐{best_count} best, ✗{reject_count} rejetés)")
    
    # Options de génération
    generation = config.get('generation', {})
    VERIFIER_MEDIAS = generation.get('verifier_medias_en_ligne', False)
    LIMITE_DEFAUT = generation.get('limite_photos_defaut', 300)
    AJOUTS_RECENTS_MIN = generation.get('ajouts_recents_minimum', 50)
    LIMITE_FAMILLE = generation.get('limite_photos_famille', 500)
    
    # Templates
    GALLERY_TEMPLATE = 'gallery_template.html'
    SPECIES_LIST_TEMPLATE = 'species_list_template.html'
    
    # Vérifier les fichiers
    if not Path(CSV_FILE).exists():
        print(f"\n❌ Fichier non trouvé: {CSV_FILE}")
        print("   Téléchargez depuis: https://ebird.org/downloadMyData")
        sys.exit(1)
    
    if not Path(TAXONOMY_FILE).exists():
        print(f"\n❌ Fichier taxonomie non trouvé: {TAXONOMY_FILE}")
        print("   Téléchargez depuis: https://ebird.org/science/use-ebird-data/the-ebird-taxonomy")
        sys.exit(1)
    
    # Initialiser le générateur
    print(f"\n📂 Chargement des données...")
    generator = EBirdGalleryGenerator(
        csv_file=CSV_FILE,
        taxonomy_file=TAXONOMY_FILE,
        media_cache_file=MEDIA_CACHE,
        traductions_file=TRADUCTIONS_FILE
    )
    
    # Charger les statuts IUCN (pour afficher dans les lightbox)
    config_menacees = config.get('especes_menacees', {})
    iucn_statuts = {}
    if config_menacees.get('activer', False):
        iucn_statuts = charger_statuts_iucn(config_menacees)
    
    # Construire le menu
    menu, familles_data = construire_menu(config, generator, curation)
    
    total_galeries = 0
    
    # ========================================
    # LISTE DES ESPÈCES
    # ========================================
    print(f"\n📋 Liste des espèces...")
    generator.generate_species_list(
        output_base='species_list',
        template_file=SPECIES_LIST_TEMPLATE,
        menu=menu,
        curation=curation,
        iucn_statuts=iucn_statuts
    )
    
    # ========================================
    # AJOUTS RÉCENTS
    # ========================================
    print(f"\n🆕 Galerie ajouts récents...")
    
    trois_mois = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
    
    # prioritize_best=False pour que les ajouts soient vraiment par date
    photos_3_mois = generator.filter_observations(
        limit=9999,
        date_start=trois_mois,
        verifier_medias_en_ligne=VERIFIER_MEDIAS,
        sort_by='date',
        curation=curation,
        prioritize_best=False
    )
    
    if len(photos_3_mois) < AJOUTS_RECENTS_MIN:
        photos_recent = generator.filter_observations(
            limit=AJOUTS_RECENTS_MIN,
            verifier_medias_en_ligne=VERIFIER_MEDIAS,
            sort_by='date',
            curation=curation,
            prioritize_best=False
        )
        print(f"   (moins de {AJOUTS_RECENTS_MIN} photos en 3 mois, affichage des {AJOUTS_RECENTS_MIN} dernières)")
    else:
        photos_recent = photos_3_mois
        print(f"   ({len(photos_recent)} photos des 3 derniers mois)")
    
    if photos_recent:
        species_set = set(p.get('scientific_name', '') for p in photos_recent)
        species_count = len(species_set)
        
        generator.generate_gallery(
            output_base='gallery_recent',
            title_fr='Ajouts récents',
            title_en='Recent Updates',
            photos=photos_recent,
            template_file=GALLERY_TEMPLATE,
            menu=menu,
            gallery_id='gallery_recent',
            species_count=species_count,
            show_date_in_overlay=True,
            iucn_statuts=iucn_statuts
        )
        total_galeries += 1
    
    # ========================================
    # MEILLEURES PHOTOS (Best of)
    # ========================================
    best_photos = generator.get_best_photos(curation=curation)
    
    if best_photos:
        print(f"\n⭐ Galerie meilleures photos...")
        species_set = set(p.get('scientific_name', '') for p in best_photos)
        species_count = len(species_set)
        
        generator.generate_gallery(
            output_base='gallery_best',
            title_fr='Meilleures photos',
            title_en='Best of',
            photos=best_photos,
            template_file=GALLERY_TEMPLATE,
            menu=menu,
            gallery_id='gallery_best',
            species_count=species_count,
            iucn_statuts=iucn_statuts
        )
        print(f"   ⭐ {len(best_photos)} photos sélectionnées ({species_count} espèces)")
        total_galeries += 1
    
    # ========================================
    # GALERIES GÉNÉRALES
    # ========================================
    galeries_generales = config.get('galeries_generales', [])
    if galeries_generales:
        print(f"\n📸 Galeries générales...")
        
        for galerie in galeries_generales:
            photos = generator.filter_observations(
                limit=galerie.get('limite', LIMITE_DEFAUT),
                countries=galerie.get('pays'),
                date_start=galerie.get('date_debut'),
                date_end=galerie.get('date_fin'),
                verifier_medias_en_ligne=VERIFIER_MEDIAS,
                curation=curation
            )
            
            if photos:
                generator.generate_gallery(
                    output_base=galerie['id'],
                    title_fr=galerie['nom_fr'],
                    title_en=galerie['nom_en'],
                    photos=photos,
                    template_file=GALLERY_TEMPLATE,
                    menu=menu,
                    gallery_id=galerie['id'],
                    iucn_statuts=iucn_statuts
                )
                total_galeries += 1
    
    # ========================================
    # VOYAGES
    # ========================================
    voyages = config.get('voyages', [])
    if voyages:
        # Trier par date_debut décroissante (plus récent d'abord)
        voyages = sorted(voyages, key=lambda v: v.get('date_debut', '0000-00-00'), reverse=True)
        
        print(f"\n✈️ Galeries voyages...")
        
        voyages_index_data = []
        
        for voyage in voyages:
            photos = generator.filter_observations(
                limit=voyage.get('limite', LIMITE_DEFAUT),
                countries=voyage.get('pays'),
                regions=voyage.get('regions'),
                date_start=voyage.get('date_debut'),
                date_end=voyage.get('date_fin'),
                verifier_medias_en_ligne=VERIFIER_MEDIAS,
                curation=curation,
                sort_by='taxonomy'
            )
            
            if photos:
                # Dates min/max
                dates = [p.get('date_raw', '') for p in photos if p.get('date_raw')]
                if dates:
                    date_min = min(dates)
                    date_max = max(dates)
                    subtitle_fr = formater_plage_dates(date_min, date_max, 'fr')
                    subtitle_en = formater_plage_dates(date_min, date_max, 'en')
                else:
                    date_min = date_max = None
                    subtitle_fr = subtitle_en = None
                
                # Frontispice: utiliser la curation (best le plus récent) ou la plus récente
                frontispice_ml = obtenir_frontispice(photos, curation)
                
                # Espèces
                species_set = set(p.get('scientific_name', '') for p in photos)
                species_count = len(species_set)
                
                # Collecter les lieux uniques pour la carte du voyage
                # Regrouper par NOM de lieu (pas par coordonnées) pour éviter les épingles multiples
                locations_seen = {}  # location_name -> location_info
                for p in photos:
                    lat = p.get('latitude')
                    lng = p.get('longitude')
                    location_name = p.get('location', '')
                    if lat and lng and lat != 0 and lng != 0 and location_name:
                        if location_name not in locations_seen:
                            locations_seen[location_name] = {
                                'lat': lat,
                                'lng': lng,
                                'name_fr': p.get('location_fr', ''),
                                'name_en': p.get('location_en', ''),
                                'location': location_name
                            }
                
                trip_locations = list(locations_seen.values()) if locations_seen else None
                
                voyages_index_data.append({
                    'nom_fr': voyage['nom_fr'],
                    'nom_en': voyage['nom_en'],
                    'file_fr': f"{voyage['id']}_fr.html",
                    'file_en': f"{voyage['id']}_en.html",
                    'frontispice_ml': frontispice_ml,
                    'dates': True if date_min else False,
                    'dates_fr': subtitle_fr,
                    'dates_en': subtitle_en,
                    'photo_count': len(photos),
                    'species_count': species_count
                })
                
                generator.generate_gallery(
                    output_base=voyage['id'],
                    title_fr=voyage['nom_fr'],
                    title_en=voyage['nom_en'],
                    photos=photos,
                    template_file=GALLERY_TEMPLATE,
                    menu=menu,
                    gallery_id=voyage['id'],
                    subtitle_fr=subtitle_fr,
                    subtitle_en=subtitle_en,
                    trip_locations=trip_locations,
                    iucn_statuts=iucn_statuts
                )
                total_galeries += 1
        
        # Page index des voyages
        if voyages_index_data:
            print(f"\n📑 Page index des voyages...")
            today = datetime.now()
            heure = today.strftime('%-Hh%M')
            update_date_fr = formater_date(today.strftime('%Y-%m-%d'), 'fr') + f' ({heure})'
            update_date_en = formater_date(today.strftime('%Y-%m-%d'), 'en') + f' ({heure})'
            
            env = Environment(loader=FileSystemLoader('.'))
            template = env.get_template('voyages_index_template.html')
            
            for lang in ['fr', 'en']:
                html = template.render(
                    lang=lang,
                    voyages=voyages_index_data,
                    menu=menu,
                    current_page=f'voyages_index_{lang}.html',
                    other_lang_page=f'voyages_index_{"en" if lang == "fr" else "fr"}.html',
                    update_date=True,
                    update_date_fr=update_date_fr,
                    update_date_en=update_date_en
                )
                with open(f'voyages_index_{lang}.html', 'w', encoding='utf-8') as f:
                    f.write(html)
            
            print(f"   ✓ voyages_index_fr.html / voyages_index_en.html ({len(voyages_index_data)} voyages)")
    
    # ========================================
    # PAYS
    # ========================================
    print(f"\n🌍 Extraction des pays...")
    
    # Extraire les pays uniques à partir des observations
    pays_stats = {}  # code_pays -> {'photos': [...], 'species': set()}
    
    for obs in generator.observations:
        state_code = obs.get('State/Province') or ''
        if not state_code:
            continue
        
        # Extraire le code pays (2 premières lettres)
        country_code = state_code.split('-')[0] if '-' in state_code else state_code
        
        if not country_code or len(country_code) != 2:
            continue
        
        ml_string = obs.get('ML Catalog Numbers') or ''
        ml_numbers = [n.strip() for n in ml_string.replace(',', ' ').split() if n.strip()]
        
        if not ml_numbers:
            continue
        
        # Vérifier que les médias sont des images
        has_image = False
        for ml in ml_numbers:
            if ml in generator.media_cache:
                if generator.media_cache[ml]['status'] == 'image':
                    has_image = True
                    break
            else:
                has_image = True  # On suppose que c'est une image si pas dans le cache
                break
        
        if not has_image:
            continue
        
        sci_name = obs.get('Scientific Name', '').strip()
        
        if country_code not in pays_stats:
            pays_stats[country_code] = {'count': 0, 'species': set()}
        
        pays_stats[country_code]['count'] += len(ml_numbers)
        if sci_name:
            pays_stats[country_code]['species'].add(sci_name)
    
    # Générer les galeries par pays
    if pays_stats:
        print(f"   Trouvé {len(pays_stats)} pays avec photos")
        
        pays_index_data = []
        total_photos_pays = 0
        
        # Traductions des pays
        traductions_pays = generator.traductions
        
        for country_code, stats in sorted(pays_stats.items(), key=lambda x: -x[1]['count']):
            # Nom du pays
            name_fr = traductions_pays.get('fr', {}).get(country_code, country_code)
            name_en = traductions_pays.get('en', {}).get(country_code, country_code)
            
            # Métadonnées du pays
            metadata = COUNTRY_METADATA.get(country_code, {
                'flag': '🏳️',
                'lat': 0,
                'lng': 0,
                'continent': 'other',
                'continent_fr': 'Autre',
                'continent_en': 'Other'
            })
            
            # Filtrer les photos pour ce pays
            photos = generator.filter_observations(
                limit=None,  # Pas de limite pour les pays
                countries=[country_code],
                verifier_medias_en_ligne=VERIFIER_MEDIAS,
                curation=curation,
                sort_by='taxonomy'
            )
            
            if photos:
                species_set = set(p.get('scientific_name', '') for p in photos)
                species_count = len(species_set)
                
                # Frontispice
                frontispice_ml = obtenir_frontispice(photos, curation)
                
                file_id = f"pays_{country_code.lower()}"
                
                pays_index_data.append({
                    'code': country_code,
                    'name_fr': name_fr,
                    'name_en': name_en,
                    'flag': metadata['flag'],
                    'lat': metadata['lat'],
                    'lng': metadata['lng'],
                    'continent': metadata['continent'],
                    'continent_fr': metadata['continent_fr'],
                    'continent_en': metadata['continent_en'],
                    'photo_count': len(photos),
                    'species_count': species_count,
                    'frontispice_ml': frontispice_ml,
                    'file_fr': f"{file_id}_fr.html",
                    'file_en': f"{file_id}_en.html"
                })
                total_photos_pays += len(photos)
                
                # Collecter les lieux pour la carte
                locations_seen = {}
                for p in photos:
                    lat = p.get('latitude')
                    lng = p.get('longitude')
                    location_name = p.get('location', '')
                    if lat and lng and lat != 0 and lng != 0 and location_name:
                        if location_name not in locations_seen:
                            locations_seen[location_name] = {
                                'lat': lat,
                                'lng': lng,
                                'name_fr': p.get('location_fr', ''),
                                'name_en': p.get('location_en', ''),
                                'location': location_name
                            }
                
                trip_locations = list(locations_seen.values()) if locations_seen else None
                
                # Générer la galerie du pays
                generator.generate_gallery(
                    output_base=file_id,
                    title_fr=name_fr,
                    title_en=name_en,
                    photos=photos,
                    template_file=GALLERY_TEMPLATE,
                    menu=menu,
                    gallery_id=file_id,
                    species_count=species_count,
                    trip_locations=trip_locations,
                    back_link_fr='pays_index_fr.html',
                    back_link_en='pays_index_en.html',
                    back_text_fr='Tous les pays',
                    back_text_en='All countries',
                    iucn_statuts=iucn_statuts
                )
                total_galeries += 1
        
        # Trier par ordre alphabétique (nom français, sans tenir compte des accents)
        def normalize_for_sort(text):
            """Enlève les accents pour un tri alphabétique correct"""
            return ''.join(
                c for c in unicodedata.normalize('NFD', text.lower())
                if unicodedata.category(c) != 'Mn'
            )
        
        pays_index_data.sort(key=lambda x: normalize_for_sort(x['name_fr']))
        
        # Page index des pays
        if pays_index_data:
            print(f"\n📑 Page index des pays...")
            today = datetime.now()
            heure = today.strftime('%-Hh%M')
            update_date_fr = formater_date(today.strftime('%Y-%m-%d'), 'fr') + f' ({heure})'
            update_date_en = formater_date(today.strftime('%Y-%m-%d'), 'en') + f' ({heure})'
            
            env = Environment(loader=FileSystemLoader('.'))
            template = env.get_template('pays_index_template.html')
            
            for lang in ['fr', 'en']:
                html = template.render(
                    lang=lang,
                    countries=pays_index_data,
                    total_countries=len(pays_index_data),
                    total_photos=total_photos_pays,
                    menu=menu,
                    current_page=f'pays_index_{lang}.html',
                    other_lang_page=f'pays_index_{"en" if lang == "fr" else "fr"}.html',
                    update_date=True,
                    update_date_fr=update_date_fr,
                    update_date_en=update_date_en
                )
                with open(f'pays_index_{lang}.html', 'w', encoding='utf-8') as f:
                    f.write(html)
            
            print(f"   ✓ pays_index_fr.html / pays_index_en.html ({len(pays_index_data)} pays)")
    
    # ========================================
    # FAMILLES
    # ========================================
    if familles_data:
        print(f"\n🪶 Galeries par famille...")

        noms_reserves = {'gallery_recent', 'gallery_best'}
        noms_reserves.update(g['id'] for g in config.get('galeries_generales', []))
        noms_reserves.update(v['id'] for v in config.get('voyages', []))

        familles_generees = 0
        for famille in familles_data:
            if famille['output_base'] in noms_reserves:
                print(f"   ⚠️  Ignoré {famille['family_code']}: collision de nom de fichier ({famille['output_base']})")
                continue

            photos = generator.filter_observations(
                limit=LIMITE_FAMILLE,
                family=famille['family_code'],
                verifier_medias_en_ligne=VERIFIER_MEDIAS,
                sort_by='taxonomy',
                curation=curation
            )

            if photos:
                species_set = set(p.get('scientific_name', '') for p in photos)
                species_count = len(species_set)

                generator.generate_gallery(
                    output_base=famille['output_base'],
                    title_fr=famille['family_code'],
                    title_en=famille['family_code'],
                    photos=photos,
                    template_file=GALLERY_TEMPLATE,
                    menu=menu,
                    gallery_id=famille['output_base'],
                    subtitle_fr=famille['desc_fr'],
                    subtitle_en=famille['desc_en'],
                    species_count=species_count,
                    back_link_fr='species_list_fr.html',
                    back_link_en='species_list_en.html',
                    back_text_fr='Liste des espèces',
                    back_text_en='Species list',
                    iucn_statuts=iucn_statuts
                )
                familles_generees += 1
                total_galeries += 1

        print(f"   ✓ {familles_generees} galeries de famille générées")

    # ========================================
    # ESPÈCES MENACÉES
    # ========================================
    config_menacees = config.get('especes_menacees', {})
    if config_menacees.get('activer', False):
        print(f"\n🔴 Galerie des espèces menacées...")
        
        # Charger les statuts IUCN
        statuts_iucn = charger_statuts_iucn(config_menacees)
        
        if statuts_iucn:
            # Mapping des statuts configuré dans YAML
            status_mapping = config_menacees.get('mapping_statuts', {
                'CR': {'name_fr': 'En danger critique', 'name_en': 'Critically Endangered', 'order': 1},
                'CR (PE)': {'name_fr': 'En danger critique (possiblement éteint)', 'name_en': 'Critically Endangered (Possibly Extinct)', 'order': 2},
                'CR (PEW)': {'name_fr': 'En danger critique (possiblement éteint à l\'état sauvage)', 'name_en': 'Critically Endangered (Possibly Extinct in the Wild)', 'order': 3},
                'EN': {'name_fr': 'En danger', 'name_en': 'Endangered', 'order': 4},
                'VU': {'name_fr': 'Vulnérable', 'name_en': 'Vulnerable', 'order': 5},
                'NT': {'name_fr': 'Quasi menacé', 'name_en': 'Near Threatened', 'order': 6},
                'DD': {'name_fr': 'Données insuffisantes', 'name_en': 'Data Deficient', 'order': 7},
                'EW': {'name_fr': 'Éteint à l\'état sauvage', 'name_en': 'Extinct in the Wild', 'order': 8},
                'EX': {'name_fr': 'Éteint', 'name_en': 'Extinct', 'order': 9},
            })
            
            # Ajouter l'ordre aux statuts
            for code, info in status_mapping.items():
                if 'order' not in info:
                    info['order'] = 99
            
            # Statuts à exclure (par défaut LC = Least Concern)
            statuts_exclus = config_menacees.get('statuts_exclus', ['LC', 'NE'])
            
            # Récupérer toutes les photos
            all_photos = generator.filter_observations(
                limit=None,
                verifier_medias_en_ligne=VERIFIER_MEDIAS,
                sort_by='taxonomy',
                curation=curation
            )
            
            # Grouper par espèce et filtrer les menacées
            species_photos = {}
            for photo in all_photos:
                sci_name = photo.get('scientific_name', '')
                if not sci_name:
                    continue
                
                # Vérifier le statut IUCN
                statut = statuts_iucn.get(sci_name)
                if not statut or statut in statuts_exclus:
                    continue
                
                # Normaliser le statut (ex: "CR (PE)" -> "CR")
                statut_normalise = statut.split()[0] if ' ' in statut else statut
                if statut not in status_mapping:
                    # Utiliser le statut normalisé s'il existe
                    if statut_normalise in status_mapping:
                        statut = statut_normalise
                    else:
                        continue
                
                if sci_name not in species_photos:
                    species_photos[sci_name] = {
                        'sci_name': sci_name,
                        'common_name_fr': photo.get('common_name_fr', sci_name),
                        'common_name_en': photo.get('common_name_en', sci_name),
                        'status': statut,
                        'taxon_order': photo.get('taxon_order', 999999),
                        'photos': []
                    }
                species_photos[sci_name]['photos'].append(photo)
            
            if species_photos:
                # Organiser par statut
                statuses_data = {}
                for sci_name, sp_data in species_photos.items():
                    statut = sp_data['status']
                    if statut not in statuses_data:
                        status_info = status_mapping.get(statut, {'name_fr': statut, 'name_en': statut, 'order': 99})
                        statuses_data[statut] = {
                            'code': statut,
                            'name_fr': status_info['name_fr'],
                            'name_en': status_info['name_en'],
                            'order': status_info['order'],
                            'species': []
                        }
                    
                    # Première photo comme représentative
                    first_photo = sp_data['photos'][0]
                    
                    statuses_data[statut]['species'].append({
                        'sci_name': sci_name,
                        'common_name_fr': sp_data['common_name_fr'],
                        'common_name_en': sp_data['common_name_en'],
                        'ml_catalog_number': first_photo.get('ml_catalog_number', ''),
                        'photo_count': len(sp_data['photos']),
                        'taxon_order': sp_data['taxon_order'],
                        'all_photos': sp_data['photos']
                    })
                
                # Trier les statuts par ordre de gravité
                statuses_list = sorted(statuses_data.values(), key=lambda x: x['order'])
                
                # Trier les espèces dans chaque statut par ordre phylogénétique
                for status in statuses_list:
                    status['species'].sort(key=lambda x: x['taxon_order'])
                
                # Compter les totaux
                total_species = len(species_photos)
                total_photos = sum(len(sp['photos']) for sp in species_photos.values())
                
                print(f"   📊 {total_species} espèces menacées, {total_photos} photos")
                
                # Générer la page
                env = Environment(loader=FileSystemLoader('.'))
                template = env.get_template('menacees_template.html')
                
                for lang in ['fr', 'en']:
                    html = template.render(
                        lang=lang,
                        statuses=statuses_list,
                        status_mapping=status_mapping,
                        total_species=total_species,
                        total_photos=total_photos,
                        menu=menu,
                        current_page=f'menacees_{lang}.html',
                        other_lang_page=f'menacees_{"en" if lang == "fr" else "fr"}.html'
                    )
                    with open(f'menacees_{lang}.html', 'w', encoding='utf-8') as f:
                        f.write(html)

                print(f"   ✓ menacees_fr.html / menacees_en.html")

                # Feed plein écran couvrant toutes les espèces menacées, par ordre de gravité de statut
                feed_photos = []
                for status in statuses_list:
                    for sp in status['species']:
                        for photo in sp['all_photos']:
                            feed_photos.append({
                                'ml_catalog_number': photo.get('ml_catalog_number', ''),
                                'common_name_fr': sp['common_name_fr'],
                                'common_name_en': sp['common_name_en'],
                                'scientific_name': sp['sci_name'],
                                'location': photo.get('location', ''),
                                'location_full_fr': photo.get('location_full_fr', ''),
                                'location_full_en': photo.get('location_full_en', ''),
                                'date_fr': photo.get('date_fr', ''),
                                'date_en': photo.get('date_en', ''),
                                'latitude': photo.get('latitude'),
                                'longitude': photo.get('longitude'),
                                'checklist_id': photo.get('checklist_id', ''),
                                'iucn_status': status['code']
                            })

                feed_template = env.get_template('gallery_feed_template.html')
                ecrire_feed_json(feed_photos, 'menacees_feed_data.json')
                with open('menacees_feed_fr.html', 'w', encoding='utf-8') as f:
                    f.write(feed_template.render(
                        lang='fr', gallery_title='Espèces menacées', photos=feed_photos, back_url='menacees_fr.html',
                        data_url='menacees_feed_data.json'
                    ))
                with open('menacees_feed_en.html', 'w', encoding='utf-8') as f:
                    f.write(feed_template.render(
                        lang='en', gallery_title='Endangered species', photos=feed_photos, back_url='menacees_en.html',
                        data_url='menacees_feed_data.json'
                    ))
                print(f"   ✓ menacees_feed_fr.html / menacees_feed_en.html (feed plein écran, {len(feed_photos)} photos)")
    
    # ========================================
    # GALERIE DES SONS
    # ========================================
    has_sounds = any(v.get('status') == 'son' for v in generator.media_cache.values())
    if has_sounds:
        print(f"\n🎵 Galerie des sons...")
        generator.generate_sounds_gallery(
            output_base='sounds_gallery',
            template_file='sounds_template.html',
            menu=menu,
            curation=curation
        )
        total_galeries += 1

    # ========================================
    # PAGE ADMIN (numéros ML)
    # ========================================
    print(f"\n🔧 Page admin...")
    generer_page_admin(config)
    
    # ========================================
    # RÉSUMÉ
    # ========================================
    print("\n" + "=" * 60)
    print("✅ GÉNÉRATION TERMINÉE")
    print(f"   📊 {total_galeries} galeries générées")
    print("=" * 60)


# ============================================================================
# PAGE ADMIN
# ============================================================================

def charger_curation(fichier_curation: str) -> dict:
    """Charge le fichier de curation des photos"""
    curation = {}
    if not Path(fichier_curation).exists():
        return curation
    
    try:
        with open(fichier_curation, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                ml = row.get('ml_number', '').strip()
                status = row.get('status', 'include').strip()
                if ml:
                    curation[ml] = status
    except Exception as e:
        print(f"   ⚠ Erreur lecture curation: {e}")
    
    return curation


def generer_page_admin(config):
    """Génère la page admin avec toutes les photos et numéros ML"""
    fichiers = config.get('fichiers', {})
    CSV_FILE = fichiers.get('donnees_ebird', 'MyEBirdData.csv')
    CACHE_FILE = fichiers.get('cache_medias', 'media_cache.csv')
    TAXONOMY_FILE = fichiers.get('taxonomie', 'eBird_taxonomy_v2025.csv')
    CURATION_FILE = fichiers.get('curation', 'photo_curation.csv')
    
    # Charger le cache des médias
    media_cache = charger_cache(CACHE_FILE)
    
    # Charger la curation existante
    curation = charger_curation(CURATION_FILE)
    if curation:
        print(f"   📋 Curation chargée: {len(curation)} photos")
    
    # Charger la taxonomie pour les noms
    taxonomy = {}
    if Path(TAXONOMY_FILE).exists():
        with open(TAXONOMY_FILE, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                sci_name = normaliser_nom_scientifique(row.get('SCI_NAME', ''))
                taxonomy[sci_name] = {
                    'common_name_en': normaliser_nom_commun(row.get('PRIMARY_COM_NAME', '')),
                    'taxon_order': int(row.get('TAXON_ORDER', 999999))
                }

    # Lire toutes les observations avec photos
    all_photos = []
    with open(CSV_FILE, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            ml_string = row.get('ML Catalog Numbers') or ''
            ml_string = ml_string.strip()
            if not ml_string:
                continue

            # Nettoyer les noms dès la lecture : enlever tout segment entre parenthèses
            # (nom commun) ou crochets (nom scientifique), ex. "Cormoran impérial
            # (groupe atriceps)" -> "Cormoran impérial"
            scientific_name = normaliser_nom_scientifique(row.get('Scientific Name', ''))
            common_name_fr = normaliser_nom_commun(row.get('Common Name', ''))
            common_name_en = taxonomy.get(scientific_name, {}).get('common_name_en', common_name_fr)
            date_raw = row.get('Date', '')
            
            for ml in ml_string.replace(',', ' ').split():
                ml = ml.strip()
                if ml:
                    # Ne garder que les images
                    media_status = media_cache.get(ml, {}).get('status', 'unknown')
                    if media_status != 'image':
                        continue
                    
                    # Déterminer le statut de curation
                    status = curation.get(ml, 'include')
                    
                    all_photos.append({
                        'ml_catalog_number': ml,
                        'scientific_name': scientific_name,
                        'common_name_fr': common_name_fr,
                        'common_name_en': common_name_en,
                        'date_raw': date_raw,
                        'status': status
                    })
    
    # Trier par date décroissante
    all_photos.sort(key=lambda x: x.get('date_raw', ''), reverse=True)
    
    # Générer la page
    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template('admin_template.html')
    
    html = template.render(photos=all_photos)
    
    with open('admin_photos.html', 'w', encoding='utf-8') as f:
        f.write(html)
    
    # Stats
    photo_best = sum(1 for p in all_photos if p['status'] == 'best')
    photo_include = sum(1 for p in all_photos if p['status'] == 'include')
    photo_reject = sum(1 for p in all_photos if p['status'] == 'reject')
    
    print(f"   ✓ admin_photos.html")
    print(f"     📷 {len(all_photos)} photos: ⭐{photo_best} best, ✓{photo_include} inclus, ✗{photo_reject} rejetés")


# ============================================================================
# VÉRIFICATION DES MÉDIAS
# ============================================================================

def verifier_medias():
    """Vérifie tous les médias et met à jour le cache"""
    config = charger_config()
    fichiers = config.get('fichiers', {})
    
    CSV_FILE = fichiers.get('donnees_ebird', 'MyEBirdData.csv')
    MEDIA_CACHE = fichiers.get('cache_medias', 'media_cache.csv')
    
    print("=" * 60)
    print("🔍 VÉRIFICATION DES MÉDIAS")
    print("=" * 60)
    
    verifier_tous_les_medias(CSV_FILE, MEDIA_CACHE)


# ============================================================================
# POINT D'ENTRÉE
# ============================================================================

def afficher_aide():
    """Affiche l'aide"""
    print("""
🐦 Générateur de galeries eBird v2.0

Usage: python generer_tout.py [commande]

Commandes:
  (aucune)    Génère toutes les galeries
  verifier    Vérifie tous les médias et met à jour le cache
  help        Affiche cette aide

Configuration: config.yaml

Fichiers requis:
  - MyEBirdData.csv (export eBird)
  - eBird_taxonomy_v2025.csv (taxonomie)
  - config.yaml (configuration)

Étapes:
  1. Téléchargez vos données depuis https://ebird.org/downloadMyData
  2. Modifiez config.yaml selon vos besoins
  3. Exécutez: python generer_tout.py verifier (première fois)
  4. Exécutez: python generer_tout.py

Options:
  python generer_tout.py           - Génère toutes les galeries
  python generer_tout.py admin     - Génère uniquement la page admin (rapide)
  python generer_tout.py verifier  - Vérifie les médias en ligne
""")


def generer_admin_seulement():
    """Génère uniquement la page admin pour la curation des photos"""
    print("=" * 60)
    print("🔧 GÉNÉRATION PAGE ADMIN UNIQUEMENT")
    print("=" * 60)
    
    config = charger_config()
    generer_page_admin(config)
    
    print("\n✅ Page admin générée: admin.html")
    print("   Ouvrez ce fichier dans votre navigateur pour gérer vos photos.")


if __name__ == '__main__':
    if len(sys.argv) > 1:
        commande = sys.argv[1].lower()
        if commande == 'verifier':
            verifier_medias()
        elif commande == 'admin':
            generer_admin_seulement()
        elif commande == 'help' or commande == '--help' or commande == '-h':
            afficher_aide()
        else:
            print(f"❌ Commande inconnue: {commande}")
            afficher_aide()
    else:
        generer_toutes_galeries()
